"""Tests for ExportService."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from datetime import datetime
import pytest
from ui_entities.inventory_item import InventoryItem, GroupedInventoryItem
from export_service import ExportService


def _make_item(name="Laptop", sub="X1", qty=3, sn=None, loc="Warehouse", cond="Good"):
    return InventoryItem(
        id=1, item_type_id=1, item_type_name=name, item_sub_type=sub,
        is_serialized=sn is not None, quantity=qty, serial_number=sn,
        location_id=1, location_name=loc, condition=cond, details="",
        created_at=datetime.now(), updated_at=datetime.now(),
    )


def _make_grouped(name="Laptop", sub="X1", qty=6, serials=None, loc="Warehouse"):
    serials = serials or []
    return GroupedInventoryItem(
        item_type_id=1, item_type_name=name, item_sub_type=sub,
        is_serialized=bool(serials), details="", total_quantity=qty,
        item_count=len(serials) or 1, serial_numbers=serials, item_ids=[1],
        location_id=1, location_name=loc,
    )


def test_build_workbook_returns_workbook():
    wb = ExportService.build_workbook([_make_item()], location_name="Warehouse")
    import openpyxl
    assert isinstance(wb, openpyxl.Workbook)


def test_items_sheet_exists():
    wb = ExportService.build_workbook([_make_item()], location_name="Warehouse")
    assert "Речі" in wb.sheetnames


def test_items_sheet_has_bold_header():
    wb = ExportService.build_workbook([_make_item()], location_name="Warehouse")
    ws = wb["Речі"]
    assert ws.cell(1, 1).font.bold is True


def test_items_sheet_header_columns():
    wb = ExportService.build_workbook([_make_item()], location_name="Warehouse")
    ws = wb["Речі"]
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == ["Тип", "Підтип", "Кількість", "Серійний номер", "Склад"]


def test_non_serialized_item_one_row():
    item = _make_grouped("Laptop", qty=5)
    wb = ExportService.build_workbook([item], location_name="Warehouse")
    ws = wb["Речі"]
    # Row 1 = header, row 2 = data
    assert ws.max_row == 2
    assert ws.cell(2, 3).value == 5  # quantity


def test_serialized_item_one_row_per_serial():
    item = _make_grouped("Laptop", serials=["SN1", "SN2", "SN3"])
    wb = ExportService.build_workbook([item], location_name="Warehouse")
    ws = wb["Речі"]
    assert ws.max_row == 4  # 1 header + 3 serials


def test_no_transactions_sheet_by_default():
    wb = ExportService.build_workbook([_make_item()], location_name="Warehouse")
    assert "Транзакції" not in wb.sheetnames


def _make_transaction(type_id=1, tx_type="add", qty_before=0, qty_after=5,
                      notes="init", sn=None, loc_id=1, from_id=None, to_id=None):
    from datetime import timezone
    return {
        "id": 1, "item_type_id": type_id, "type": tx_type,
        "quantity_change": qty_after - qty_before,
        "quantity_before": qty_before, "quantity_after": qty_after,
        "notes": notes, "serial_number": sn,
        "location_id": loc_id, "from_location_id": from_id, "to_location_id": to_id,
        "created_at": datetime(2026, 2, 24, 10, 30, tzinfo=timezone.utc),
    }


def test_transactions_sheet_created_when_passed():
    wb = ExportService.build_workbook(
        [_make_item()], "Warehouse",
        transactions=[_make_transaction()],
        loc_map={1: "Warehouse"},
        type_map={1: "Laptop \u2014 X1"},
    )
    assert "Транзакції" in wb.sheetnames


def test_transactions_sheet_has_bold_header():
    wb = ExportService.build_workbook(
        [_make_item()], "Warehouse",
        transactions=[_make_transaction()],
    )
    ws = wb["Транзакції"]
    assert ws.cell(1, 1).font.bold is True


def test_transactions_sheet_header_columns():
    wb = ExportService.build_workbook(
        [_make_item()], "Warehouse",
        transactions=[_make_transaction()],
    )
    ws = wb["Транзакції"]
    headers = [ws.cell(1, c).value for c in range(1, 13)]
    assert headers == [
        "Дата", "Транзакція", "Тип", "Підтип", "Серійний номер", "Кількість",
        "Кількість до", "Кількість після", "Нотатки", "Склад", "Зі складу", "На склад",
    ]


def test_transactions_sheet_data_row():
    wb = ExportService.build_workbook(
        [_make_item()], "Warehouse",
        transactions=[_make_transaction(type_id=1, tx_type="add", qty_before=0, qty_after=5)],
        loc_map={1: "Warehouse"},
        type_map={1: "Laptop \u2014 X1"},
    )
    ws = wb["Транзакції"]
    assert ws.cell(2, 2).value == "ADD"
    assert ws.cell(2, 3).value == "Laptop"
    assert ws.cell(2, 4).value == "X1"
    assert ws.cell(2, 6).value == 5   # quantity_change = 5-0
    assert ws.cell(2, 7).value == 0   # qty_before
    assert ws.cell(2, 8).value == 5   # qty_after


def test_transactions_location_names_resolved():
    tx = _make_transaction(tx_type="transfer", from_id=1, to_id=2)
    wb = ExportService.build_workbook(
        [_make_item()], "Warehouse",
        transactions=[tx],
        loc_map={1: "Warehouse A", 2: "Warehouse B"},
    )
    ws = wb["Транзакції"]
    assert ws.cell(2, 11).value == "Warehouse A"   # Зі складу
    assert ws.cell(2, 12).value == "Warehouse B"   # На склад
