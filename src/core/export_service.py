"""Excel export service — builds openpyxl workbooks from inventory data."""

from __future__ import annotations

from typing import Dict, List, Optional, Union

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ui.models.inventory_item import GroupedInventoryItem, InventoryItem
from ui.translations import format_quantity_change


class ExportService:
    """Builds Excel workbooks from inventory data. No UI dependencies."""

    # Sheet column headers
    ITEM_HEADERS = ["Тип", "Підтип", "Кількість", "Серійний номер", "Склад"]
    TRANSACTION_HEADERS = [
        "Дата", "Транзакція", "Тип", "Підтип", "Серійний номер", "Кількість",
        "Кількість до", "Кількість після", "Нотатки", "Склад", "Зі складу", "На склад",
    ]

    @staticmethod
    def build_workbook(
        items: List[Union[InventoryItem, GroupedInventoryItem]],
        location_name: str,
        transactions: Optional[List[dict]] = None,
        loc_map: Optional[Dict[int, str]] = None,
        type_map: Optional[Dict[int, str]] = None,
    ) -> openpyxl.Workbook:
        """Build a workbook with an Items sheet and optional Transactions sheet.

        Args:
            items: List of InventoryItem or GroupedInventoryItem for the Items sheet.
                Location context is embedded per-row via each item's location_name field.
            location_name: Accepted for API compatibility; not embedded in the workbook
                (sheets are named "Майно"/"Транзакції", and each row carries
                its own location value).
            transactions: Optional list of transaction dicts (from _transaction_to_dict).
            loc_map: Optional {location_id: name} for resolving location IDs in transactions.
            type_map: Optional {item_type_id: display_name} for transaction sheet.

        Returns:
            openpyxl.Workbook ready to be saved.
        """
        wb = openpyxl.Workbook()
        ws_items = wb.active
        ws_items.title = "Майно"

        ExportService._write_items_sheet(ws_items, items)

        if transactions is not None:
            ws_tx = wb.create_sheet("Транзакції")
            ExportService._write_transactions_sheet(
                ws_tx, transactions, loc_map or {}, type_map or {}
            )

        return wb

    @staticmethod
    def _write_items_sheet(
        ws: Worksheet,
        items: List[Union[InventoryItem, GroupedInventoryItem]],
    ) -> None:
        bold = Font(bold=True)

        # Header row
        for col, header in enumerate(ExportService.ITEM_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        row = 2
        for item in items:
            if isinstance(item, GroupedInventoryItem) and item.is_serialized:
                # One row per serial number
                for sn in item.serial_numbers:
                    ws.cell(row=row, column=1, value=item.item_type_name)
                    ws.cell(row=row, column=2, value=item.item_sub_type or "")
                    ws.cell(row=row, column=3, value=1)
                    ws.cell(row=row, column=4, value=sn)
                    ws.cell(row=row, column=5, value=item.location_name)
                    row += 1
            elif isinstance(item, GroupedInventoryItem):
                ws.cell(row=row, column=1, value=item.item_type_name)
                ws.cell(row=row, column=2, value=item.item_sub_type or "")
                ws.cell(row=row, column=3, value=item.total_quantity)
                ws.cell(row=row, column=4, value="")
                ws.cell(row=row, column=5, value=item.location_name)
                row += 1
            else:
                # InventoryItem
                ws.cell(row=row, column=1, value=item.item_type_name)
                ws.cell(row=row, column=2, value=item.item_sub_type or "")
                ws.cell(row=row, column=3, value=item.quantity)
                ws.cell(row=row, column=4, value=item.serial_number or "")
                ws.cell(row=row, column=5, value=item.location_name)
                row += 1

        ExportService._autofit(ws, len(ExportService.ITEM_HEADERS))

    @staticmethod
    def _write_transactions_sheet(
        ws: Worksheet,
        transactions: List[dict],
        loc_map: Dict[int, str],
        type_map: Dict[int, str],
    ) -> None:
        bold = Font(bold=True)

        for col, header in enumerate(ExportService.TRANSACTION_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        for row, trans in enumerate(transactions, start=2):
            created = trans.get("created_at")
            date_str = created.strftime("%d.%m.%Y %H:%M") if created else ""
            type_name = trans.get("type", "").upper()
            item_type_id = trans.get("item_type_id")
            name_parts = (type_map.get(item_type_id) or str(item_type_id)).split(" \u2014 ", 1)
            item_name = name_parts[0]
            item_sub = name_parts[1] if len(name_parts) > 1 else ""
            from_id = trans.get("from_location_id")
            to_id = trans.get("to_location_id")
            location_id = trans.get("location_id")

            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=type_name)
            ws.cell(row=row, column=3, value=item_name)
            ws.cell(row=row, column=4, value=item_sub)
            ws.cell(row=row, column=5, value=trans.get("serial_number") or "")
            ws.cell(row=row, column=6, value=format_quantity_change(trans))
            ws.cell(row=row, column=7, value=trans.get("quantity_before", ""))
            ws.cell(row=row, column=8, value=trans.get("quantity_after", ""))
            ws.cell(row=row, column=9, value=trans.get("notes") or "")
            ws.cell(row=row, column=10, value=loc_map.get(location_id, "") if location_id else "")
            ws.cell(row=row, column=11, value=loc_map.get(from_id, "") if from_id else "")
            ws.cell(row=row, column=12, value=loc_map.get(to_id, "") if to_id else "")

        ExportService._autofit(ws, len(ExportService.TRANSACTION_HEADERS))

    @staticmethod
    def _autofit(ws, num_cols: int, min_width: int = 10, max_width: int = 50) -> None:
        """Set column widths based on content."""
        for col in range(1, num_cols + 1):
            max_len = 0
            for cell in ws[get_column_letter(col)]:
                try:
                    max_len = max(max_len, len(str(cell.value if cell.value is not None else "")))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col)].width = max(
                min_width, min(max_len + 2, max_width)
            )
